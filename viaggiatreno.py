
from datetime import datetime
from urllib.request import urlopen 
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
import json 
#import xlsxwriter
#from xlsxwriter.utility import xl_col_to_name

import os.path


#hasmap numerotreno:codicestazione
treni = {
    "5838" : "S09211",
    "5834" : "S09211"
}
# current date and time
dataOdierna = datetime.now()
dataOdierna = dataOdierna.replace(hour=0, minute=0, second=0, microsecond=0)
#oggi = dataCompleta.date() 
print('Date and Time is:', dataOdierna)
timestamp = int(round(datetime.timestamp(dataOdierna)))
nomeFile = 'report_orari_treno_'+str(dataOdierna.year)+str(dataOdierna.month)+'.xlsx'
path = './'+nomeFile
checkFile = os.path.isfile(path)
print("Verifica esistenza del file :" + str(checkFile))
if checkFile==True:
    wb = load_workbook(nomeFile)
    print ("Il file " + nomeFile + " esiste")
else:
    wb = Workbook()
    print ("Il file " + nomeFile + " non esiste")
if 'Sheet' in wb.sheetnames:
    wb.remove(wb['Sheet'])
for codiceTreno in treni:
    try:
        url = "http://www.viaggiatreno.it/infomobilita/resteasy/viaggiatreno/andamentoTreno/"+treni[codiceTreno]+"/"+codiceTreno+"/"+str(timestamp)+"000"
        print("url =", url)
        response = urlopen(url) 
        print(response.getcode())
        nomeSheet = codiceTreno #+'_'+data_json['origine'][0:3]+'_'+data_json['destinazione'][0:3]+'_'+orarioDiPartenza.strftime("%H")+orarioDiPartenza.strftime("%M")
        if nomeSheet in wb.sheetnames:
            ws = wb[nomeSheet]
            print ("Lo sheet " + nomeSheet + " esiste")
        else:
            ws = wb.create_sheet(nomeSheet)
            print ("Lo sheet " + nomeSheet + " non esiste")
        rowCell = dataOdierna.day+1
        colCell = 1
        print("data odierna =", dataOdierna)
        ws.cell(row=1, column=1).value = "Giorno"
        ws.cell(row=34, column=1).value = "Totale Ritardo"
        ws.cell(row=35, column=1).value = "Media Ritardo"
        ws.cell(row=rowCell, column=colCell).value = dataOdierna
        if response.getcode()==200:
            data_json = json.loads(response.read())   
            orarioDiPartenza = datetime.fromtimestamp(int(round(data_json['orarioPartenzaZero']/1000)))
            print("%s:%s" % (orarioDiPartenza.strftime("%H"), orarioDiPartenza.strftime("%M")))
            colCell=2
            for fermate in data_json['fermate']:
                print(str(fermate['stazione']) + "," + str(fermate['ritardo']))
                ws.cell(row=1, column=colCell).value = str(fermate['stazione'])
                if (fermate['ritardo']>0):
                    ws.cell(row=rowCell, column=colCell).value = fermate['ritardo']
                    ws.cell(row=rowCell, column=colCell).fill = PatternFill(start_color='fde910',end_color='fde910',fill_type='solid')
                    print('yellow')
                else:
                    ws.cell(row=rowCell, column=colCell).value = fermate['ritardo']
                columnLetter= get_column_letter(colCell)
                ws.cell(row=34, column=colCell).value = '=SUM(%s2:%s32)' %(columnLetter,columnLetter)
                ws.cell(row=35, column=colCell).value = '=AVERAGE(%s2:%s32)' %(columnLetter,columnLetter)
                colCell = colCell+1
        else:
            ws.cell(row=rowCell, column=2).value = "N/D"
            print("Informazioni non disponibili per il treno "  +treni[codiceTreno]+ " e per il giorno " +str(dataOdierna))
    except:
        print("Si è verificato un errore nel recupero delle informazioni") 
wb.save(nomeFile)